#!/usr/bin/env python3
"""
Reads CLAS, Fixture and the full per-day predictions/standings from all Excel
files matching ADMINExcelMundial2026*.xlsx and writes a merged docs/data.json.

The daily data covers EVERY day of the tournament (not just the one selected in
the DailyPrediction dropdown): it is read from the ADMIN master table, which
holds every match for every day plus each player's prediction and points. The
web page then offers a day selector to browse any date.

To support more than 25 players, simply add a second file named e.g.
  ADMINExcelMundial2026_2.xlsx
The script will discover it automatically, merge all players, and
re-rank the combined standings (overall and per day).
"""

import glob
import json
import os
import warnings
from collections import OrderedDict
from datetime import datetime, timezone, timedelta

warnings.filterwarnings("ignore", category=UserWarning)
import openpyxl

REPO_ROOT   = os.path.join(os.path.dirname(__file__), "..")
OUTPUT_PATH = os.path.join(REPO_ROOT, "docs", "data.json")
# Where to look for the Excel files; CI overrides this to point at
# LibreOffice-recalculated temp copies that never get committed.
EXCEL_DIR   = os.environ.get("EXCEL_DIR", REPO_ROOT)

MAX_VALUES = {
    "f_grupos":   432,
    "pos_grupos": 156,
    "eq_16":       96,
    "pt_16":      144,
    "eq_8":        64,
    "pt_8":        96,
    "eq_4":        40,
    "pt_4":        60,
    "eq_2":        28,
    "pt_2":        36,
    "eq_34":       10,
    "eq_final":    20,
    "pt_34":       18,
    "pt_final":    21,
    "honor":       71,
}

COLUMNS = [
    ("pos",         "Pos"),
    ("jugador",     "Jugador"),
    ("puntos",      "Puntos Totales"),
    ("f_grupos",    "F. Grupos"),
    ("pos_grupos",  "Pos. Grupos"),
    ("eq_16",       "Equipos 1/16"),
    ("pt_16",       "Partidos 1/16"),
    ("eq_8",        "Equipos 1/8"),
    ("pt_8",        "Partidos 1/8"),
    ("eq_4",        "Equipos 1/4"),
    ("pt_4",        "Partidos 1/4"),
    ("eq_2",        "Equipos 1/2"),
    ("pt_2",        "Partidos 1/2"),
    ("eq_34",       "Equipos 3-4"),
    ("eq_final",    "Equipos Final"),
    ("pt_34",       "Partido 3-4"),
    ("pt_final",    "Partido Final"),
    ("honor",       "Cuadro de Honor"),
]

PHASE_MAP = {
    range(1, 73):    "grupos",
    range(73, 89):   "dieciseisavos",
    range(89, 97):   "octavos",
    range(97, 101):  "cuartos",
    range(101, 103): "semis",
    range(103, 104): "tercero",
    range(104, 105): "final",
}

PHASE_LABELS = {
    "grupos":        "Fase de Grupos",
    "dieciseisavos": "1/16 de Final",
    "octavos":       "1/8 de Final",
    "cuartos":       "Cuartos de Final",
    "semis":         "Semifinales",
    "tercero":       "Tercer Puesto",
    "final":         "Final",
}

# Names that identify unfilled template placeholder rows in a second file.
PLACEHOLDER_PREFIXES = ("pegar valores", "pegar nombre", "jugador ")


def is_placeholder(name):
    return isinstance(name, str) and name.lower().strip().startswith(PLACEHOLDER_PREFIXES)


def to_num(val):
    if val is None or val == "-":
        return 0
    try:
        return int(val)
    except (ValueError, TypeError):
        return 0


def fmt_date(dt):
    if dt is None:
        return None
    if hasattr(dt, "strftime"):
        return dt.strftime("%Y-%m-%d %H:%M")
    return str(dt)


def get_phase(num):
    if not isinstance(num, int):
        return "final"
    for rng, phase in PHASE_MAP.items():
        if num in rng:
            return phase
    return "final"


def assign_positions(players, sort_key="puntos"):
    """Sort players by sort_key descending and assign positions (ties share the same rank)."""
    players.sort(key=lambda p: p.get(sort_key, 0), reverse=True)
    pos = 1
    for i, p in enumerate(players):
        if i > 0 and p.get(sort_key, 0) < players[i - 1].get(sort_key, 0):
            pos = i + 1
        p["pos"] = pos
    return players


# ── CLAS ─────────────────────────────────────────────────────────────────────

def extract_clas_players(ws):
    """Return (title, [players]) from a CLAS worksheet."""
    rows = list(ws.iter_rows(values_only=True))
    title = rows[0][1] or "CLASIFICACIÓN MUNDIAL 2026"
    players = []
    for row in rows[4:]:
        if not isinstance(row[2], str) or row[2] == "-" or is_placeholder(row[2]):
            continue
        players.append({
            # .strip() para evitar nombres con espacios sobrantes (p. ej.
            # "jaidali ") que romperían la unión por nombre con el bloque diario,
            # que sí viene recortado.
            "jugador":    row[2].strip(),
            "puntos":     to_num(row[3]),
            "f_grupos":   to_num(row[4]),
            "pos_grupos": to_num(row[5]),
            "eq_16":      to_num(row[6]),
            "pt_16":      to_num(row[7]),
            "eq_8":       to_num(row[8]),
            "pt_8":       to_num(row[9]),
            "eq_4":       to_num(row[10]),
            "pt_4":       to_num(row[11]),
            "eq_2":       to_num(row[12]),
            "pt_2":       to_num(row[13]),
            "eq_34":      to_num(row[14]),
            "eq_final":   to_num(row[15]),
            "pt_34":      to_num(row[16]),
            "pt_final":   to_num(row[17]),
            "honor":      to_num(row[18]),
        })
    return title, players


# ── FIXTURE ───────────────────────────────────────────────────────────────────

def extract_fixture(ws):
    matches = []
    current_group = None
    for row in ws.iter_rows(values_only=True):
        grp = row[35]
        if grp and isinstance(grp, str) and grp.startswith("Grupo"):
            current_group = grp

        date = row[23]
        home = row[26]
        away = row[31]
        num  = row[33]
        if not (date and hasattr(date, "strftime") and home and away):
            continue

        phase  = get_phase(num)
        goal_h = row[28]
        goal_a = row[29]
        matches.append({
            "num":         num,
            "date":        fmt_date(date),
            "matchday":    str(row[25]) if row[25] else None,
            "group":       current_group if phase == "grupos" else None,
            "home":        str(home),
            "away":        str(away),
            "goal_home":   int(goal_h) if isinstance(goal_h, (int, float)) else None,
            "goal_away":   int(goal_a) if isinstance(goal_a, (int, float)) else None,
            "phase":       phase,
            "phase_label": PHASE_LABELS.get(phase, phase),
        })

    matches.sort(key=lambda m: (m["date"] or "", m["num"] or 999))
    return matches


# ── GROUP STANDINGS (tablas de posiciones, hoja WORLDCUP) ─────────────────────
#
# La hoja WORLDCUP trae un bloque "display" ya ordenado por posición. Cada grupo
# ocupa 8 filas y las columnas (0-based) son:
#   35 Pos | 37 Selección | 38 Pts | 39 J | 40 G | 41 E | 42 P | 43 GF | 44 GC | 45 DG
# La etiqueta "Grupo X" está en la fila idx 3+8g (col 35), el encabezado en 4+8g
# y las 4 selecciones en 5+8g..8+8g. Ese bloque ya aplica el desempate del Excel
# (incluyendo orden de seed y puntos penalizados), así que solo se transcribe.

GS_LABEL_COL = 35
GS_COLS = {
    "pos": 35, "team": 37, "pts": 38, "pj": 39, "g": 40,
    "e": 41, "p": 42, "gf": 43, "gc": 44, "dg": 45,
}


def extract_group_standings(ws):
    rows = list(ws.iter_rows(values_only=True))
    groups = []
    g = 0
    while True:
        lbl_idx = 3 + 8 * g
        if lbl_idx >= len(rows):
            break
        lbl_row = rows[lbl_idx]
        label = lbl_row[GS_LABEL_COL] if len(lbl_row) > GS_LABEL_COL else None
        if not (isinstance(label, str) and label.strip().lower().startswith("grupo")):
            break
        teams = []
        for t in range(4):
            ri = 5 + 8 * g + t
            if ri >= len(rows):
                break
            row = rows[ri]
            team = row[GS_COLS["team"]] if len(row) > GS_COLS["team"] else None
            if not (isinstance(team, str) and team.strip()):
                continue
            teams.append({k: (team.strip() if k == "team" else to_num(row[col]))
                          for k, col in GS_COLS.items()})
        groups.append({"name": label.strip(), "teams": teams})
        g += 1
    return groups


# ── DAILY (ALL DAYS, from the ADMIN master table) ─────────────────────────────
#
# The DailyPrediction / DailyClas worksheets only show ONE day (driven by a
# dropdown), so reading them with data_only=True yields a single day. The ADMIN
# sheet instead holds the full master table: every match for every day, plus
# each player's prediction and points. We read that to expose ALL days online.
#
# Layout of the ADMIN sheet:
#   row index 4 (5th row) is the header; player names live at columns
#     18, 21, 24, ...  (stride 3). For player at column C, that player's
#     prediction is at column C and the awarded points at column C+1.
#   match rows start at index 5; for each:
#     col 6  = "Ref fecha"  (integer day group, same value = same day)
#     col 7  = "Fecha"      (a datetime for calendar days, or a phase label
#                            string such as "Octavos de final" / "3-4 & Final")
#     col 10 = match name   (e.g. "México-Sudáfrica")

DAILY_PLAYER_START = 18    # first player column
DAILY_PLAYER_STRIDE = 3    # columns between players
DAILY_HEADER_ROW = 4       # 0-based index of the player-name header row
DAILY_FIRST_MATCH_ROW = 5
DAILY_REF_COL = 6
DAILY_DATE_COL = 7
DAILY_MATCH_COL = 10


def extract_daily_all(ws):
    """Return an ordered list of day dicts from an ADMIN worksheet.

    Each day dict: {ref, label, date, matches, predictions, clas}
      - label   : human label ("11/06/2026" or a phase name)
      - date    : ISO date string ("2026-06-11") or None for phase groups
      - matches : [match name, ...]
      - predictions : [{jugador, preds:[str|None, ...]}, ...]
      - clas        : [{jugador, puntos_dia, pts:[int, ...]}, ...]
    """
    rows = list(ws.iter_rows(values_only=True))
    header = rows[DAILY_HEADER_ROW]

    player_cols = []
    c = DAILY_PLAYER_START
    while c < len(header):
        name = header[c]
        if isinstance(name, str) and name.strip() and not is_placeholder(name):
            player_cols.append((c, name.strip()))
        c += DAILY_PLAYER_STRIDE

    # Group match rows by their "Ref fecha" while preserving sheet order.
    days = OrderedDict()
    for row in rows[DAILY_FIRST_MATCH_ROW:]:
        name = row[DAILY_MATCH_COL]
        date = row[DAILY_DATE_COL]
        ref = row[DAILY_REF_COL]
        if not (isinstance(name, str) and name.strip() and date is not None and ref is not None):
            continue
        if ref not in days:
            if hasattr(date, "strftime"):
                label = date.strftime("%d/%m/%Y")
                iso = date.strftime("%Y-%m-%d")
            else:
                label = str(date).strip()
                iso = None
            days[ref] = {"ref": ref, "label": label, "date": iso, "rows": []}
        days[ref]["rows"].append((str(name).strip(), row))

    result = []
    for ref, d in days.items():
        matches = [m[0] for m in d["rows"]]
        predictions, clas = [], []
        for col, pname in player_cols:
            preds, pts = [], []
            for _, row in d["rows"]:
                pred = row[col] if col < len(row) else None
                point = row[col + 1] if col + 1 < len(row) else None
                preds.append(None if (pred is None or pred == "-") else str(pred).strip())
                pts.append(to_num(point))
            predictions.append({"jugador": pname, "preds": preds})
            clas.append({"jugador": pname, "puntos_dia": sum(pts), "pts": pts})
        result.append({
            "ref": ref,
            "label": d["label"],
            "date": d["date"],
            "matches": matches,
            "predictions": predictions,
            "clas": clas,
        })
    return result


# ── MAIN ──────────────────────────────────────────────────────────────────────

def main():
    # Discover all Excel files (sorted so primary always comes first)
    pattern = os.path.join(EXCEL_DIR, "ADMINExcelMundial2026*.xlsx")
    excel_files = sorted(glob.glob(pattern))

    if not excel_files:
        raise FileNotFoundError(f"No Excel files found matching {pattern}")

    # First file alphabetically is the primary (provides Fixture + match columns)
    primary_file = excel_files[0]
    print(f"Found {len(excel_files)} Excel file(s): {[os.path.basename(f) for f in excel_files]}")
    print(f"Primary file: {os.path.basename(primary_file)}")

    vet     = timezone(timedelta(hours=-4))
    now_vet = datetime.now(vet)

    # ── Merge CLAS and the full per-day daily data from all files ────────────
    all_clas_players = []
    title = "CLASIFICACIÓN MUNDIAL 2026"
    fixture = []
    groups = []
    # Per-day merge keyed by day "ref"; the primary file defines the day order,
    # labels and match list, and every file appends its own players.
    merged_days = OrderedDict()

    for path in excel_files:
        wb = openpyxl.load_workbook(path, data_only=True)
        is_primary = (os.path.abspath(path) == os.path.abspath(primary_file))

        t, players = extract_clas_players(wb["CLAS"])
        if is_primary:
            title = t
        all_clas_players.extend(players)

        for day in extract_daily_all(wb["ADMIN"]):
            ref = day["ref"]
            if ref not in merged_days:
                merged_days[ref] = {
                    "ref": ref,
                    "label": day["label"],
                    "date": day["date"],
                    "matches": day["matches"],
                    "predictions": [],
                    "clas": [],
                }
            merged_days[ref]["predictions"].extend(day["predictions"])
            merged_days[ref]["clas"].extend(day["clas"])

        # Fixture y tablas de posiciones solo desde el archivo primario
        if is_primary:
            fixture = extract_fixture(wb["WORLDCUP"])
            groups = extract_group_standings(wb["WORLDCUP"])

    # ── Re-rank merged standings (overall + per day) ──────────────────────────
    clas_players = assign_positions(all_clas_players, "puntos")

    daily = []
    for day in merged_days.values():
        assign_positions(day["clas"], "puntos_dia")
        daily.append(day)

    output = {
        "updated_at": now_vet.strftime("%d/%m/%Y %H:%M (hora Venezuela)"),
        "clas": {"title": title, "players": clas_players},
        "max_values":  MAX_VALUES,
        "columns":     COLUMNS,
        "fixture":     fixture,
        "groups":      groups,
        "daily":       daily,
    }

    os.makedirs(os.path.dirname(OUTPUT_PATH), exist_ok=True)
    with open(OUTPUT_PATH, "w", encoding="utf-8") as f:
        json.dump(output, f, ensure_ascii=False, indent=2)

    print(f"Wrote {len(clas_players)} players, {len(fixture)} matches → {OUTPUT_PATH}")


if __name__ == "__main__":
    main()
